import pandas as pd
import xlrd
from openpyxl import load_workbook
from io import BytesIO
import os

def read_excel_file(file_stream, page, limit):
    file_bytes = file_stream.read()
    try:
        # Try with openpyxl (for .xlsx)
        return read_xlsx(BytesIO(file_bytes), page, limit)
    except Exception as e:
        print("Failed .xlsx, trying .xls", e)
        try:
            return read_xls(BytesIO(file_bytes), page, limit)
        except Exception as ex:
            return {"error": f"Unable to read Excel file: {str(ex)}"}

def read_xlsx(buffer, page, limit):
    wb = load_workbook(filename=buffer, read_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    start = (page - 1) * limit
    end = start + limit
    paginated = data_rows[start:end]

    # Convert headers to str
    result = [dict(zip([str(h) for h in headers], row)) for row in paginated]

    return {
        "file_type": "excel",
        "total": len(data_rows),
        "page": page,
        "limit": limit,
        "data": result
    }



def read_xls(buffer, page, limit):
    wb = xlrd.open_workbook(file_contents=buffer.read())
    sheet = wb.sheet_by_index(0)

    headers = sheet.row_values(0)
    rows = [sheet.row_values(row_idx) for row_idx in range(1, sheet.nrows)]

    start = (page - 1) * limit
    end = start + limit
    paginated = rows[start:end]

    # Convert all header keys to string
    result = [dict(zip([str(h) for h in headers], row)) for row in paginated]

    return {
        "file_type": "excel",
        "total": len(rows),
        "page": page,
        "limit": limit,
        "data": result
    }

